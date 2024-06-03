from os import mkdir
from openpyxl import load_workbook

def dlya_3_groups():
    file_name = 'D:\\NIR\\LIST_NEW\\first_course.xlsx'
    wb = load_workbook(file_name)
    sheet = wb['Поток М2']

    sheet.insert_cols(6)
    sheet.insert_cols(6)
    sheet.insert_cols(13)
    sheet.insert_cols(13)

    sheet['E2'].value = sheet['B1'].value
    sheet['F2'].value = sheet['D1'].value
    sheet['L2'].value = sheet['I1'].value
    sheet['M2'].value = sheet['K1'].value
    sheet['S2'].value = sheet['P1'].value
    sheet['T2'].value = sheet['R1'].value

    sheet.delete_rows(1)

    number_of_values_1 = sum(1 for cell in sheet['B'] if cell.value is not None)
    number_of_values_2 = sum(1 for cell in sheet['J'] if cell.value is not None)
    number_of_values_3 = sum(1 for cell in sheet['Q'] if cell.value is not None)


    for row1 in range(2, number_of_values_1 + 1):
        sheet[f'E{row1}'] = sheet[f'E{row1 - 1}'].value
        sheet[f'F{row1}'] = sheet[f'F{row1 - 1}'].value


    for row2 in range(2, number_of_values_2 + 1):
        sheet[f'L{row2}'] = sheet[f'L{row2 - 1}'].value
        sheet[f'M{row2}'] = sheet[f'M{row2 - 1}'].value


    for row3 in range(2, number_of_values_3 + 1):
        sheet[f'S{row3}'] = sheet[f'S{row3 - 1}'].value
        sheet[f'T{row3}'] = sheet[f'T{row3 - 1}'].value

    source_start_row = 1
    source_start_col = 8
    destination_start_row = number_of_values_1 + 1
    destination_start_col = 1
    number_of_rows = number_of_values_2
    number_of_cols = 6

    for row_2 in range(number_of_rows):
        for col_2 in range(number_of_cols):
            value = sheet.cell(row=source_start_row + row_2, column=source_start_col + col_2).value
            sheet.cell(row=destination_start_row + row_2, column=destination_start_col + col_2).value = value

    destination_start_row += number_of_rows
    wb.save(file_name)

dlya_3_groups()
