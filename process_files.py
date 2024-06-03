##
import os
import win32com.client as client
import shutil

"""def delete_files_in_folder(folder):
    for filename in os.listdir(folder):
        file_path = os.path.join(folder, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))"""
def process_files():
    folder = 'D:\\NIR\\LIST_NEW'
    excel = client.Dispatch("Excel.Application")

    for file in os.listdir("D:\\NIR\\LIST\\"):
        filename, file_extension = os.path.splitext(file)
        wb = excel.Workbooks.Open(os.path.join("D:\\NIR\\LIST\\", file))
        output_path = os.path.join("D:\\NIR\\LIST_NEW\\", filename + ".xlsx")
        wb.SaveAs(output_path, 51)

        wb.Close()
process_files()

def move_files():
    source_folder = "D:\\NIR\\LIST_NEW"
    destination_folder = "D:\\NIR\\LIST_NEW"

    for file in os.listdir(source_folder):
        if os.path.isfile(os.path.join(source_folder, file)):  # Проверка, является ли объект файлом
            filename, file_extension = os.path.splitext(file)
            if file_extension == ".xlsx":
                course_name = filename.split('_')[0]  # Получаем название курса из имени файла
                course_folder = os.path.join(destination_folder, course_name + "_courses")  # Папка курса
                if not os.path.exists(course_folder):
                    os.makedirs(course_folder)  # Создаем папку, если она не существует
                shutil.move(os.path.join(source_folder, file), os.path.join(course_folder, file))  # Перемещаем файл

move_files()

##
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import pandas as pd
def work_with_file():
    file_path = r'D:\NIR\LIST_NEW\first_course.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['Поток М1']

    # Вставка новых пустых столбцов F и G
    sheet.insert_cols(6)
    sheet.insert_cols(6)

    # Вставка новых пустых столбцов M и N
    sheet.insert_cols(13)
    sheet.insert_cols(13)

    # Копирование значений
    sheet['E2'].value = sheet['B1'].value
    sheet['F2'].value = sheet['D1'].value
    sheet['L2'].value = sheet['I1'].value
    sheet['M2'].value = sheet['K1'].value
    sheet['S2'].value = sheet['P1'].value
    sheet['U2'].value = sheet['R1'].value
    sheet['T2'].value = sheet['U2'].value
    sheet['U2'].value = None

    # Заполнение формулами
    for row in range(3, 34):
        sheet[f'E{row}'] = f"=E{row - 1}"
        sheet[f'F{row}'] = f"=F{row - 1}"
        sheet[f'L{row}'] = f"=L{row - 1}"
        sheet[f'M{row}'] = f"=M{row - 1}"
        sheet[f'S{row}'] = f"=S{row - 1}"
        sheet[f'T{row}'] = f"=T{row - 1}"

    # Копирование значений и вставка их как значений
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = cell.value

    # Копирование диапазонов
    copy_range(sheet, 'H1:M33', 'A34')
    copy_range(sheet, 'O1:T33', 'A67')

    # Очистка диапазонов
    clear_range(sheet, 'H1:AB46')

    # Установка заголовков
    headers = ["№", "Имя", "Фамилия", "Отчество", "Группа", "Направление"]
    for i, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=i)
        cell.value = header
        cell.font = Font(bold=True)

    # Автофит для столбцов A:G
    for col in range(1, 8):
        max_length = 0
        column = get_column_letter(col)
        for cell in sheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    sheet.auto_filter.ref = sheet.dimensions
    filter_data(file_path)

    workbook.save(file_path)




def copy_range(sheet, source_range, target_start):
    # Копирование значений из одного диапазона в другой
    source_cells = sheet[source_range]
    start_row = int(target_start[1:])
    start_col = openpyxl.utils.cell.column_index_from_string(target_start[0])

    for i, row in enumerate(source_cells):
        for j, cell in enumerate(row):
            new_cell = sheet.cell(row=start_row + i, column=start_col + j)
            new_cell.value = cell.value


def clear_range(sheet, cell_range):
    for row in sheet[cell_range]:
        for cell in row:
            cell.value = None

def filter_data(file_path):
    df = pd.read_excel(file_path, sheet_name='Поток И1 (1,2,3)')
    df2 = df[(df['№'] != "null")]
    return df2

work_with_file()