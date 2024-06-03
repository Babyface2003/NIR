from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Protection, PatternFill
import time

def dlya_3_groups():
    file_name = 'D:\\NIR\\LIST_NEW\\mag_first_course.xlsx'
    wb = load_workbook(file_name)
    sheet = wb['12.04.01']

    sheet.insert_cols(6)
    sheet.insert_cols(6)
    sheet.insert_cols(13)
    sheet.insert_cols(13)

    # Copy value and style
    def copy_value_and_style(source_cell, target_cell):
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom,
                diagonal=source_cell.border.diagonal,
                diagonal_direction=source_cell.border.diagonal_direction,
                outline=source_cell.border.outline,
                vertical=source_cell.border.vertical,
                horizontal=source_cell.border.horizontal
            )
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
            target_cell.number_format = source_cell.number_format
            target_cell.protection = Protection(
                locked=source_cell.protection.locked,
                hidden=source_cell.protection.hidden
            )
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                textRotation=source_cell.alignment.textRotation,
                wrapText=source_cell.alignment.wrapText,
                shrinkToFit=source_cell.alignment.shrinkToFit,
                indent=source_cell.alignment.indent
            )

    copy_value_and_style(sheet['B1'], sheet['E2'])
    copy_value_and_style(sheet['D1'], sheet['F2'])
    copy_value_and_style(sheet['I1'], sheet['L2'])
    copy_value_and_style(sheet['K1'], sheet['M2'])
    copy_value_and_style(sheet['P1'], sheet['S2'])
    copy_value_and_style(sheet['R1'], sheet['T2'])

    sheet.delete_rows(1)

    number_of_values_1 = sum(1 for cell in sheet['B'] if cell.value is not None)
    number_of_values_2 = sum(1 for cell in sheet['J'] if cell.value is not None)
    number_of_values_3 = sum(1 for cell in sheet['Q'] if cell.value is not None)

    for row1 in range(2, number_of_values_1 + 1):
        copy_value_and_style(sheet[f'E{row1 - 1}'], sheet[f'E{row1}'])
        copy_value_and_style(sheet[f'F{row1 - 1}'], sheet[f'F{row1}'])

    for row2 in range(2, number_of_values_2 + 1):
        copy_value_and_style(sheet[f'L{row2 - 1}'], sheet[f'L{row2}'])
        copy_value_and_style(sheet[f'M{row2 - 1}'], sheet[f'M{row2}'])

    for row3 in range(2, number_of_values_3 + 1):
        copy_value_and_style(sheet[f'S{row3 - 1}'], sheet[f'S{row3}'])
        copy_value_and_style(sheet[f'T{row3 - 1}'], sheet[f'T{row3}'])

    time.sleep(2)

    source_start_row = 1
    source_start_col = 8
    destination_start_row = number_of_values_1 + 1
    destination_start_col = 1
    number_of_rows = number_of_values_2
    number_of_cols = 6

    for row_2 in range(number_of_rows):
        for col_2 in range(number_of_cols):
            source_cell = sheet.cell(row=source_start_row + row_2, column=source_start_col + col_2)
            target_cell = sheet.cell(row=destination_start_row + row_2, column=destination_start_col + col_2)
            copy_value_and_style(source_cell, target_cell)

    destination_start_row += number_of_rows

    time.sleep(2)

    source_start_row_2 = 1
    source_start_col_2 = 15
    number_of_rows_2 = number_of_values_3
    number_of_cols_2 = 6

    for row_3 in range(number_of_rows_2):
        for col_3 in range(number_of_cols_2):
            source_cell = sheet.cell(row=source_start_row_2 + row_3, column=source_start_col_2 + col_3)
            target_cell = sheet.cell(row=destination_start_row + row_3, column=destination_start_col + col_3)
            copy_value_and_style(source_cell, target_cell)

    wb.save(file_name)

dlya_3_groups()
